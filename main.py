import rhinoscriptsyntax as rs
import clr
clr.AddReference("RhinoCommon")
import Rhino
import Rhino.Geometry
import os

from openpyxl import load_workbook   # Module needed for loading existing workbook
from openpyxl.drawing.image import Image

rs.EnableRedraw(False)

NamaLayer=[] #siapkan list nama2 layer
#Quantity=[]
layers=rs.LayerNames() # nama2 semua layer
#masukkan nama2 layer yang aktif ke list NamaLayer
for layer in layers:
    if rs.IsLayerOn(layer):  # filter hanya pada layer yg on saja
        rs.ObjectsByLayer(layer, True)  # select object pada layer yang on
        objs = rs.SelectedObjects()
        if objs:  # jika layer ini aja object nya, daftarkan nama layer ini kedalam list NamaLayer
            NamaLayer.append(layer)
        rs.UnselectAllObjects()
print(NamaLayer)

layertobom = []
for layer in layers:
    nama = layer
    cari = ':'
    index = nama.rfind(cari)
    if index >= 0:
        #slice disini
        index = index+1
        layertobom.append(nama[index:len(nama)])
    else:
        layertobom.append(nama)
print(layertobom)

os.system('copy BOM_Blank.xlsx BOM.xlsx')
wb = load_workbook(filename="BOM.xlsx")
ws = wb.active

img = Image('image.jpg'
            '')
ws.add_image(img, 'A1')

for index, value in enumerate(layertobom):
    cell = ws.cell(row=index+9, column=2)
    cell.value = value

# Save the spreadsheet
wb.save(filename="BOM.xlsx")





"""
https://stevebaer.wordpress.com/
-----------------------------------------------
Returns the object type
    Parameters:
      object_id (guid): identifier of an object
    Returns:
      number: The object type if successful.
        The valid object types are as follows:
        Value   Description
          0           Unknown object
          1           Point
          2           Point cloud
          4           Curve
          8           Surface or single-face brep
          16          Polysurface or multiple-face
          32          Mesh
          256         Light
          512         Annotation
          4096        Instance or block reference
          8192        Text dot object
          16384       Grip object
          32768       Detail
          65536       Hatch
          131072      Morph control
          134217728   Cage
          268435456   Phantom
          536870912   Clipping plane
          1073741824  Extrusion
    Example:
      import rhinoscriptsyntax as rs
      obj = rs.GetObject("Select object")
      if obj:
          objtype = rs.ObjectType(obj)
          print "Object type:", objtype

-----------------------------------------------


Hi all,

I'm trying to script Rhino -a 3d CAD nurbs modeler- via python.
Which looks to be very cool, if it wasn't for some gnarly problem I'm
running into:

from jf.interface import RhinoScript #interface to rhino by makepy
import os, sys, win32com.client

Rhino = win32com.client.Dispatch("Rhino3.Application")
time.sleep(1)
Rhino.Visible = True
RS = Rhino.GetScriptObject
RS = RhinoScript.IRxRhino(RS)

So far so good.
Now if I use RS.AddLine((0,0,0),(10,10,10)) -which produces a line from 2
coords- all works well.
Too bad that when I have to use an array of points all goes wrong:
RS.Curve(points,degree)
RS.AddCurve(((1,2,3),(4,5,6),(7,8,9)),3)
---
Traceback (most recent call last):
  File "<input>", line 1, in ?
  File "c:\Python23\lib\site-packages\jf\interface\RhinoScript.py", line 48,
in AddCurve
    return self._ApplyTypes_(77, 1, (12, 0), ((12, 0), (12, 16)),
'AddCurve', None,vaPoints, vaDegree)
  File "c:\Python23\lib\site-packages\win32com\client\__init__.py", line
446, in _ApplyTypes_
    return self._get_good_object_(
com_error: (-2147352567, 'Exception occurred.', (6, 'RhinoScript_m', 'Type
mismatch in parameter.  One-dimensional array required.', 'C:\\Program
Files\\Rhinoceros 3.0\\RhinoScript_m.HLP', 393222, 0), None)
---


Even when I use it own method of getting points -via the GUI- I get the same
error:
RS.AddCurve(RS.GetPoints(),3)
---
Traceback (most recent call last):
  File "<input>", line 1, in ?
  File "c:\Python23\lib\site-packages\jf\interface\RhinoScript.py", line 48,
in AddCurve
    return self._ApplyTypes_(77, 1, (12, 0), ((12, 0), (12, 16)),
'AddCurve', None,vaPoints, vaDegree)
  File "c:\Python23\lib\site-packages\win32com\client\__init__.py", line
446, in _ApplyTypes_
    return self._get_good_object_(
com_error: (-2147352567, 'Exception occurred.', (6, 'RhinoScript_m', 'Type
mismatch in parameter.  One-dimensional array required.', 'C:\\Program
Files\\Rhinoceros 3.0\\RhinoScript_m.HLP', 393222, 0), None)
---

More in detail:
s = RS.GetPoints()
s
((-7.0, -30.0, 0.0), (15.0, -24.0, 0.0), (-7.0, 12.0, 0.0), (14.0, 29.0,
0.0), (28.0, 10.0, 0.0), (20.0, 1.0, 0.0))

Cool, I get to choose points in the GUI, right back into python, wicked!
Though when I send these points back to the app:
RS.AddPolyline(s)
The same error all over again:
---
Traceback (most recent call last):
  File "<input>", line 1, in ?
  File "c:\Python23\lib\site-packages\jf\interface\RhinoScript.py", line
129, in AddPolyline
    return self._ApplyTypes_(85, 1, (12, 0), ((12, 0),), 'AddPolyline',
None,vaPoints)
  File "c:\Python23\lib\site-packages\win32com\client\__init__.py", line
446, in _ApplyTypes_
    return self._get_good_object_(
com_error: (-2147352567, 'Exception occurred.', (6, 'RhinoScript_m', 'Type
mismatch in parameter.  One-dimensional array required.', 'C:\\Program
Files\\Rhinoceros 3.0\\RhinoScript_m.HLP', 393222, 0), None)
---

RS.AddPolyline(((1,1,1),(2,2,2),(3,3,3)))
Again, the same error persists.

So this leaves me to believe that the problem lies in sending the array to
the app... Now I have no idea to solve this... but having too much fun to
give this up... any suggestions?

Cheers,

Jelle.

"""