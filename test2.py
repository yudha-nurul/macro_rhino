import os
from openpyxl import load_workbook   # Module needed for loading existing workbook
from openpyxl.drawing.image import Image

layers = ['induknya::Layer 02', 'Layer 04']
print(layers)
layertobom = []
for layer in layers:
    nama = layer
    cari = ':'
    index = nama.rfind(cari)
    if index >= 0:
        index = index+1
        layertobom.append(nama[index:len(nama)])
    else:
        layertobom.append(nama)
print(layertobom)

os.system('copy BOM_Blank.xlsx BOM.xlsx')
wb = load_workbook(filename="BOM.xlsx")
ws = wb.active

img = Image('image.jpg'
            '')
ws.add_image(img, 'A1')

for index, value in enumerate(layertobom):
    cell = ws.cell(row=index+9, column=2)
    cell.value = value

# Save the spreadsheet
wb.save(filename="BOM.xlsx")
